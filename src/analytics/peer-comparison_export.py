import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

DB_PATH = "data/nifty100.db"
OUTPUT_PATH = "output/peer_comparison.xlsx"

METRIC_LABELS = {
    "return_on_equity_pct_pctile": "ROE %ile",
    "return_on_capital_pct_pctile": "ROCE %ile",
    "net_profit_margin_pct_pctile": "NPM %ile",
    "debt_to_equity_pctile": "D/E %ile (lower debt = higher score)",
}

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_peer_percentiles(db_path=DB_PATH):
    conn = sqlite3.connect(db_path)
    df = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    conn.close()
    return df


def color_for_percentile(value):
    if pd.isna(value):
        return None
    if value >= 66:
        return GREEN
    elif value >= 33:
        return YELLOW
    return RED


def build_peer_comparison_workbook(df, output_path=OUTPUT_PATH):
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    metric_cols = list(METRIC_LABELS.keys())
    display_cols = ["company_id", "is_benchmark"] + metric_cols

    for peer_group_name, group_df in df.groupby("peer_group_name"):
        sheet_name = str(peer_group_name)[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        headers = ["Company", "Benchmark"] + list(METRIC_LABELS.values())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for _, row in group_df.sort_values(metric_cols[0], ascending=False).iterrows():
            row_values = [row["company_id"], "Yes" if row["is_benchmark"] else ""]
            row_values += [row[c] for c in metric_cols]
            ws.append(row_values)

            excel_row = ws.max_row
            for col_idx, metric_col in enumerate(metric_cols, start=3):
                cell = ws.cell(row=excel_row, column=col_idx)
                fill = color_for_percentile(row[metric_col])
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    df = load_peer_percentiles()
    path = build_peer_comparison_workbook(df)
    print(f"Saved {path}")
    print(f"Sheets created: {df['peer_group_name'].nunique()}")